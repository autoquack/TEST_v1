from openpyxl import load_workbook
workbook = load_workbook("C:\\Users\martin.haluska\Downloads\pythonTestFolder//hello_world_2.xlsx")
workbook.sheetnames = ['Sheet 1']

sheet = workbook.active
sheet
#Worksheet "Sheet 1">

sheet.title = "Sheet 1"