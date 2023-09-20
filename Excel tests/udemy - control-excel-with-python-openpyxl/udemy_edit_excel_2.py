from openpyxl import Workbook
wb = Workbook ()
ws = wb.create_sheet("A sheet", 0)
ws2 = wb.create_sheet("Sheet nr 2")
for sheet in wb:
    print(sheet.title)

wb.remove(wb["Sheet"])
for sheet in wb:
    print(sheet.title)

del wb["A sheet"]
for sheet in wb:
    print(sheet.title)

ws3 = wb.create_sheet("something 1")

ws.title
ws.title = "New title"
for sheet in wb:
    print(sheet.title)

wb.save("Added_sheets2.xlsx")