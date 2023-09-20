import openpyxl
from openpyxl import Workbook
#
#wb = Workbook()
wb = openpyxl.load_workbook()
ws1 = wb.create_sheet("Sheet A", 0)
ws2 = wb.create_sheet("Sheet B")
for sheet in wb:
    print(sheet.title)
