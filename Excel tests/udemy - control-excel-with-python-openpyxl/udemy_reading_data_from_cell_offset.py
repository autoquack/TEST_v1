from openpyxl import Workbook
wb = Workbook()
ws = wb.worksheets[0]

ws["A1"].value = "Train"
ws.cell(1,1).offset(0,1).value = "Train cart"
ws["A1"].value
ws["B1"].value

ws["A1"].offset(0,1).value = "Train cart 2"
ws["A1"].value, ws["B1"].value

mother_cell = ws.cell(3,3)
child_cell = mother_cell.offset(0,1)

mother_cell

child_cell

mother_cell.value = "Mother Theresa"
child_cell.value = "Da Vinci the son"
mother_cell, mother_cell.value
child_cell, child_cell.value

wb.save("Added_sheets5.xlsx")