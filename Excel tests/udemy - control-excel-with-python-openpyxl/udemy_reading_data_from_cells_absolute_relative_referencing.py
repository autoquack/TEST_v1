import openpyxl
wb = openpyxl.Workbook()
ws = wb.worksheets[0]
ws["A1"].value = 56
ws["C3"].value = "abc"
ws["A1"]

ws.cell(row=1, column=1).value = 1111
ws.cell(row=3, column=3).value = "def"
ws["C3"].value

ws.cell(3,3).value=999
ws["C3"].value

from openpyxl.utils.cell import get_column_letter
get_column_letter(3)
get_column_letter(16)
get_column_letter(1)

ws[get_column_letter(1)+"2"].value=34
ws["A2"]
ws["A2"].value


wb.save("Added_sheets4.xlsx")