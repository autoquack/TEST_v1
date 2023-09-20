from openpyxl import load_workbook

wb = load_workbook("2.4_Hello_copies.xlsx")
for sheet in wb:
    print(sheet.title)

ws1 = wb.worksheets[0]
ws2 = wb.worksheets[1]
worksheets = wb.sheetnames
type(worksheets)

print(worksheets)
ws3 = wb(worksheets[2])
for sheet in wb:
    print(sheet.title)

