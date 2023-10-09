# test

from openpyxl import Workbook, load_workbook

wb = load_workbook("testbook.xlsx")      #wb je skratka pre workbook (ako nazov premennej)
#ws = wb.active      #ws je skratka work sheet (ako nazov premennej), .active otvori aktivny sheet
ws = wb["Sheetos1"]   #otvori specificky sheet

### VYTVORI NOVY SHEET V DOKUMENTE ###

wb.create_sheet("Sheetos 2")

### VYPISE NAZOV SHEETU ###
print(ws)        #vypise mi aktivny sheet name
#print(ws["Sheetos1"]    #vypise specificky sheet

### VYTLACI VSETKY MENASHEETOV ###
print(wb.sheetnames)    #vytlaci vsetky mena sheetov

### ZAPISE HODNOTU DO BUNKY ###
ws["G9"].value = "Test"     # zapisem hodnotu Test do pola G9

### ULOZI UPRAVY ###
wb.save("testbook.xlsx")       # ulozim workbook