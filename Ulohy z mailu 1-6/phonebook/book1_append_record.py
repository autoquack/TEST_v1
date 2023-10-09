import openpyxl

#workbook = openpyxl.load_workbook("C:\Users\martin.haluska\PycharmProjects\TEST_v1\Ulohy z mailu 1-6\phonebook\phonebook.xlsx")

from openpyxl import load_workbook

#file_path = "D:\\ExcelApp\\MyCompanyStaff.xlsx"  # Setting the path to (location of) the new Excel workbook

#workbook.save(file_path)

wb_append = load_workbook("C:\\Users\\martin.haluska\\PycharmProjects\\TEST_v1\\Ulohy z mailu 1-6\\phonebook\\phonebook.xlsx")

sheet = wb_append.active

rows = (

    ("Albert", "Sobota", "0949090909"),

    ("Bertram", "Nedela", "0940929304"),

)


for row in rows:
    sheet.append(row)


wb_append.save('phonebook.xlsx')